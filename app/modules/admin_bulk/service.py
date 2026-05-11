from __future__ import annotations

import csv
from datetime import date
from io import BytesIO, StringIO
from typing import Any, Callable

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundError, ValidationError
from app.core.security import hash_password
from app.modules.academic.model import AcademicYear, Branch, Class, Course, Section, Subject
from app.modules.exams.model import Exam, ExamSubject, ExamWorkflow, Mark
from app.modules.fees.model import FeePayment, FeeStatus, FeeStructure, FeeType, StudentFee
from app.modules.library.model import Book
from app.modules.roles.model import Role, UserRole
from app.modules.students.model import Student, StudentAcademicRecord, StudentDocument, StudentStatus
from app.modules.teachers.model import Teacher
from app.modules.users.model import User


def _b(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    return str(value).strip().lower() in {"1", "yes", "y", "true", "active", "current"}


def _s(row: dict, key: str, default: str = "") -> str:
    return str(row.get(key, default) or default).strip()


def _i(row: dict, key: str, default: int = 0) -> int:
    value = _s(row, key)
    return int(float(value)) if value else default


def _f(row: dict, key: str, default: float = 0) -> float:
    value = _s(row, key)
    return float(value) if value else default


def _d(row: dict, key: str) -> date | None:
    value = _s(row, key)
    return date.fromisoformat(value) if value else None


def _csv_bytes(headers: list[str], rows: list[dict[str, Any]]) -> bytes:
    stream = StringIO()
    writer = csv.DictWriter(stream, fieldnames=headers, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({header: row.get(header, "") for header in headers})
    return stream.getvalue().encode("utf-8-sig")


def template_bytes(resource: str) -> bytes:
    spec = SPECS.get(resource)
    if not spec:
        raise NotFoundError("Unknown bulk resource")
    return _csv_bytes(spec["headers"], spec["examples"])


def parse_upload(filename: str, content: bytes) -> list[dict[str, str]]:
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValidationError("XLSX upload requires openpyxl. Use the CSV template for now.") from exc
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or "").strip() for cell in rows[0]]
        return [
            {headers[i]: "" if value is None else str(value).strip() for i, value in enumerate(row) if i < len(headers)}
            for row in rows[1:]
            if any(value not in (None, "") for value in row)
        ]

    text = content.decode("utf-8-sig")
    return [
        {key.strip(): (value or "").strip() for key, value in row.items()}
        for row in csv.DictReader(StringIO(text))
        if any((value or "").strip() for value in row.values())
    ]


async def _one(db: AsyncSession, model, *conditions):
    return (await db.execute(select(model).where(*conditions))).scalar_one_or_none()


async def _role(db: AsyncSession, institution_id: str, slug: str) -> Role | None:
    return await _one(db, Role, Role.institution_id == institution_id, Role.slug == slug)


async def _ensure_user(
    db: AsyncSession,
    institution_id: str,
    role_slug: str,
    *,
    email: str,
    username: str,
    full_name: str,
    phone: str | None = None,
    password: str = "Demo@123",
) -> User:
    user = await _one(db, User, User.email == email)
    if not user:
        user = User(
            institution_id=institution_id,
            email=email,
            username=username,
            password_hash=hash_password(password),
            full_name=full_name,
            phone=phone,
            is_active=True,
            is_superuser=False,
        )
        db.add(user)
        await db.flush()
    else:
        user.institution_id = institution_id
        user.username = username or user.username
        user.full_name = full_name or user.full_name
        user.phone = phone or user.phone
        user.is_active = True
        if password:
            user.password_hash = hash_password(password)

    role = await _role(db, institution_id, role_slug)
    if role:
        existing = await _one(db, UserRole, UserRole.user_id == user.id, UserRole.role_id == role.id)
        if not existing:
            db.add(UserRole(user_id=user.id, role_id=role.id))
    return user


async def _year(db: AsyncSession, institution_id: str, label: str) -> AcademicYear:
    item = await _one(db, AcademicYear, AcademicYear.institution_id == institution_id, AcademicYear.label == label)
    if not item:
        raise NotFoundError(f"Academic year not found: {label}")
    return item


async def _course(db: AsyncSession, institution_id: str, code: str) -> Course:
    item = await _one(db, Course, Course.institution_id == institution_id, Course.code == code)
    if not item:
        raise NotFoundError(f"Course not found: {code}")
    return item


async def _branch(db: AsyncSession, institution_id: str, code: str, course_code: str = "") -> Branch:
    query = (
        select(Branch)
        .join(Course, Course.id == Branch.course_id)
        .where(Course.institution_id == institution_id, Branch.code == code)
    )
    if course_code:
        query = query.where(Course.code == course_code)
    rows = (
        await db.execute(
            query
        )
    ).scalars().all()
    if not rows:
        raise NotFoundError(f"Branch not found: {code}")
    if len(rows) > 1:
        raise ValidationError(f"Branch code is ambiguous: {code}. Provide course_code.")
    return rows[0]


async def _class(db: AsyncSession, institution_id: str, name: str, course_code: str = "", branch_code: str = "") -> Class:
    query = (
        select(Class)
        .join(Course, Course.id == Class.course_id)
        .where(Course.institution_id == institution_id, Class.name == name)
    )
    if course_code:
        query = query.where(Course.code == course_code)
    if branch_code:
        query = query.outerjoin(Branch, Branch.id == Class.branch_id).where(Branch.code == branch_code)
    rows = (
        await db.execute(
            query
        )
    ).scalars().all()
    if not rows:
        raise NotFoundError(f"Class not found: {name}")
    if len(rows) > 1:
        raise ValidationError(f"Class name is ambiguous: {name}. Provide course_code and branch_code.")
    return rows[0]


async def _section(
    db: AsyncSession,
    institution_id: str,
    class_name: str,
    section_name: str,
    course_code: str = "",
    branch_code: str = "",
) -> Section:
    class_ = await _class(db, institution_id, class_name, course_code, branch_code)
    item = await _one(db, Section, Section.class_id == class_.id, Section.name == section_name)
    if not item:
        raise NotFoundError(f"Section not found: {class_name} / {section_name}")
    return item


async def _subject(db: AsyncSession, institution_id: str, code: str) -> Subject:
    item = (
        await db.execute(
            select(Subject)
            .join(Course, Course.id == Subject.course_id)
            .where(Course.institution_id == institution_id, Subject.code == code)
        )
    ).scalar_one_or_none()
    if not item:
        raise NotFoundError(f"Subject not found: {code}")
    return item


async def _student(db: AsyncSession, roll_number: str) -> Student:
    item = await _one(db, Student, Student.roll_number == roll_number)
    if not item:
        raise NotFoundError(f"Student not found: {roll_number}")
    return item


async def import_rows(db: AsyncSession, institution_id: str, resource: str, rows: list[dict[str, str]]) -> dict:
    handler = IMPORTERS.get(resource)
    if not handler:
        raise NotFoundError("Unknown bulk resource")

    result = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
    for index, row in enumerate(rows, start=2):
        try:
            action = await handler(db, institution_id, row)
            result[action] += 1
        except Exception as exc:
            result["errors"].append({"row": index, "error": str(exc), "data": row})
    await db.flush()
    return result


async def export_rows(db: AsyncSession, institution_id: str, resource: str) -> bytes:
    spec = SPECS.get(resource)
    exporter = EXPORTERS.get(resource)
    if not spec or not exporter:
        raise NotFoundError("Unknown bulk resource")
    rows = await exporter(db, institution_id)
    return _csv_bytes(spec["headers"], rows)


async def _import_academic_year(db, institution_id, row):
    label = _s(row, "label")
    item = await _one(db, AcademicYear, AcademicYear.institution_id == institution_id, AcademicYear.label == label)
    action = "updated" if item else "created"
    if not item:
        item = AcademicYear(institution_id=institution_id, label=label, start_date=_d(row, "start_date"), end_date=_d(row, "end_date"))
        db.add(item)
    item.start_date = _d(row, "start_date") or item.start_date
    item.end_date = _d(row, "end_date") or item.end_date
    item.is_current = _b(row.get("is_current"))
    item.is_active = _b(row.get("is_active"), True)
    return action


async def _import_course(db, institution_id, row):
    code = _s(row, "code")
    item = await _one(db, Course, Course.institution_id == institution_id, Course.code == code)
    action = "updated" if item else "created"
    if not item:
        item = Course(institution_id=institution_id, code=code, name=_s(row, "name"), level=_s(row, "level", "UG"), duration_years=_i(row, "duration_years", 3))
        db.add(item)
    item.name = _s(row, "name", item.name)
    item.level = _s(row, "level", item.level)
    item.duration_years = _i(row, "duration_years", item.duration_years)
    item.is_active = _b(row.get("is_active"), True)
    return action


async def _import_branch(db, institution_id, row):
    course = await _course(db, institution_id, _s(row, "course_code"))
    code = _s(row, "code")
    item = await _one(db, Branch, Branch.course_id == course.id, Branch.code == code)
    action = "updated" if item else "created"
    if not item:
        item = Branch(course_id=course.id, code=code, name=_s(row, "name"))
        db.add(item)
    item.name = _s(row, "name", item.name)
    item.is_active = _b(row.get("is_active"), True)
    return action


async def _import_class(db, institution_id, row):
    course = await _course(db, institution_id, _s(row, "course_code"))
    branch = await _branch(db, institution_id, _s(row, "branch_code"), _s(row, "course_code")) if _s(row, "branch_code") else None
    year = await _year(db, institution_id, _s(row, "academic_year_label")) if _s(row, "academic_year_label") else None
    name = _s(row, "name")
    item = await _one(db, Class, Class.course_id == course.id, Class.name == name)
    action = "updated" if item else "created"
    if not item:
        item = Class(course_id=course.id, branch_id=branch.id if branch else None, academic_year_id=year.id if year else None, name=name)
        db.add(item)
    item.branch_id = branch.id if branch else None
    item.academic_year_id = year.id if year else None
    item.year_no = _i(row, "year_no", item.year_no or 1)
    item.semester = _i(row, "semester", item.semester or 1)
    item.intake_capacity = _i(row, "intake_capacity", item.intake_capacity or 60)
    return action


async def _import_section(db, institution_id, row):
    class_ = await _class(db, institution_id, _s(row, "class_name"), _s(row, "course_code"), _s(row, "branch_code"))
    name = _s(row, "name")
    item = await _one(db, Section, Section.class_id == class_.id, Section.name == name)
    action = "updated" if item else "created"
    if not item:
        item = Section(class_id=class_.id, name=name)
        db.add(item)
    item.max_strength = _i(row, "max_strength", item.max_strength or 60)
    return action


async def _import_subject(db, institution_id, row):
    course = await _course(db, institution_id, _s(row, "course_code"))
    class_ = await _class(db, institution_id, _s(row, "class_name"), _s(row, "course_code"), _s(row, "branch_code"))
    branch = await _branch(db, institution_id, _s(row, "branch_code"), _s(row, "course_code")) if _s(row, "branch_code") else None
    year = await _year(db, institution_id, _s(row, "academic_year_label")) if _s(row, "academic_year_label") else None
    code = _s(row, "code")
    item = await _one(db, Subject, Subject.course_id == course.id, Subject.code == code)
    action = "updated" if item else "created"
    if not item:
        item = Subject(course_id=course.id, class_id=class_.id, code=code, name=_s(row, "name"))
        db.add(item)
    item.class_id = class_.id
    item.branch_id = branch.id if branch else None
    item.academic_year_id = year.id if year else None
    item.semester = _i(row, "semester", item.semester or 1)
    item.name = _s(row, "name", item.name)
    item.credits = _i(row, "credits", item.credits or 0)
    item.is_active = _b(row.get("is_active"), True)
    return action


async def _import_student(db, institution_id, row):
    user = await _ensure_user(
        db,
        institution_id,
        "student",
        email=_s(row, "email"),
        username=_s(row, "username") or _s(row, "email").split("@")[0],
        full_name=_s(row, "full_name"),
        phone=_s(row, "phone") or None,
        password=_s(row, "password", "Demo@123"),
    )
    roll = _s(row, "roll_number")
    item = await _one(db, Student, Student.roll_number == roll)
    action = "updated" if item else "created"
    if not item:
        item = Student(user_id=user.id, roll_number=roll)
        db.add(item)
        await db.flush()
    item.user_id = user.id
    item.date_of_birth = _d(row, "date_of_birth")
    item.gender = _s(row, "gender") or None
    item.guardian_name = _s(row, "guardian_name") or None
    item.guardian_phone = _s(row, "guardian_phone") or None
    item.guardian_email = _s(row, "guardian_email") or None
    if item.guardian_email:
        await _ensure_user(
            db,
            institution_id,
            "parent",
            email=item.guardian_email,
            username=f"parent.{item.guardian_email.split('@')[0]}",
            full_name=item.guardian_name or f"Parent of {user.full_name}",
            phone=item.guardian_phone,
            password="Demo@123",
        )

    year = await _year(db, institution_id, _s(row, "academic_year_label"))
    branch = await _branch(db, institution_id, _s(row, "branch_code"), _s(row, "course_code"))
    section = await _section(db, institution_id, _s(row, "class_name"), _s(row, "section_name"), _s(row, "course_code"), _s(row, "branch_code"))
    active = await _one(db, StudentAcademicRecord, StudentAcademicRecord.student_id == item.id, StudentAcademicRecord.exited_at == None)
    if not active:
        db.add(StudentAcademicRecord(student_id=item.id, section_id=section.id, branch_id=branch.id, academic_year_id=year.id, status=StudentStatus.ACTIVE))
    else:
        active.section_id = section.id
        active.branch_id = branch.id
        active.academic_year_id = year.id
        active.status = StudentStatus.ACTIVE
    return action


async def _import_teacher(db, institution_id, row):
    user = await _ensure_user(
        db,
        institution_id,
        "teacher",
        email=_s(row, "email"),
        username=_s(row, "username") or _s(row, "email").split("@")[0],
        full_name=_s(row, "full_name"),
        phone=_s(row, "phone") or None,
        password=_s(row, "password", "Demo@123"),
    )
    code = _s(row, "employee_code")
    item = await _one(db, Teacher, Teacher.employee_code == code)
    action = "updated" if item else "created"
    if not item:
        item = Teacher(user_id=user.id, employee_code=code)
        db.add(item)
    item.user_id = user.id
    item.designation = _s(row, "designation") or None
    item.joined_at = _d(row, "joined_at")
    return action


async def _import_fee_type(db, institution_id, row):
    name = _s(row, "name")
    item = await _one(db, FeeType, FeeType.institution_id == institution_id, FeeType.name == name)
    action = "updated" if item else "created"
    if not item:
        item = FeeType(institution_id=institution_id, name=name)
        db.add(item)
    item.description = _s(row, "description") or None
    return action


async def _import_fee_structure(db, institution_id, row):
    fee_type = await _one(db, FeeType, FeeType.institution_id == institution_id, FeeType.name == _s(row, "fee_type_name"))
    if not fee_type:
        raise NotFoundError(f"Fee type not found: {_s(row, 'fee_type_name')}")
    course = await _course(db, institution_id, _s(row, "course_code"))
    year = await _year(db, institution_id, _s(row, "academic_year_label"))
    item = await _one(db, FeeStructure, FeeStructure.fee_type_id == fee_type.id, FeeStructure.course_id == course.id, FeeStructure.academic_year_id == year.id)
    action = "updated" if item else "created"
    if not item:
        item = FeeStructure(fee_type_id=fee_type.id, course_id=course.id, academic_year_id=year.id, amount=_f(row, "amount"), frequency=_s(row, "frequency", "annual"))
        db.add(item)
    item.amount = _f(row, "amount", item.amount)
    item.frequency = _s(row, "frequency", item.frequency)
    return action


async def _import_student_fee(db, institution_id, row):
    student = await _student(db, _s(row, "roll_number"))
    fee_type = await _one(db, FeeType, FeeType.institution_id == institution_id, FeeType.name == _s(row, "fee_type_name"))
    course = await _course(db, institution_id, _s(row, "course_code"))
    year = await _year(db, institution_id, _s(row, "academic_year_label"))
    structure = await _one(db, FeeStructure, FeeStructure.fee_type_id == fee_type.id, FeeStructure.course_id == course.id, FeeStructure.academic_year_id == year.id)
    if not structure:
        raise NotFoundError("Fee structure not found for selected fee type/course/year")
    item = await _one(db, StudentFee, StudentFee.student_id == student.id, StudentFee.fee_structure_id == structure.id)
    action = "updated" if item else "created"
    if not item:
        item = StudentFee(student_id=student.id, fee_structure_id=structure.id, amount_due=_f(row, "amount_due") or structure.amount)
        db.add(item)
    item.amount_due = _f(row, "amount_due", item.amount_due)
    item.due_date = _d(row, "due_date")
    item.status = FeeStatus.PAID if float(item.amount_paid or 0) >= float(item.amount_due or 0) else FeeStatus.PARTIAL if float(item.amount_paid or 0) else FeeStatus.UNPAID
    return action


async def _import_student_document(db, institution_id, row):
    student = await _student(db, _s(row, "roll_number"))
    title = _s(row, "title")
    item = await _one(db, StudentDocument, StudentDocument.student_id == student.id, StudentDocument.title == title)
    action = "updated" if item else "created"
    if not item:
        item = StudentDocument(student_id=student.id, title=title, document_type=_s(row, "document_type", "general"))
        db.add(item)
    item.document_type = _s(row, "document_type", item.document_type)
    item.file_name = _s(row, "file_name") or None
    item.file_url = _s(row, "file_url") or None
    item.status = _s(row, "status", item.status or "pending")
    item.remarks = _s(row, "remarks") or None
    return action


async def _import_exam(db, institution_id, row):
    year = await _year(db, institution_id, _s(row, "academic_year_label"))
    name = _s(row, "name")
    item = await _one(db, Exam, Exam.institution_id == institution_id, Exam.academic_year_id == year.id, Exam.name == name)
    action = "updated" if item else "created"
    if not item:
        item = Exam(institution_id=institution_id, academic_year_id=year.id, name=name, exam_type=_s(row, "exam_type", "midterm"))
        db.add(item)
    item.exam_type = _s(row, "exam_type", item.exam_type)
    item.workflow_status = ExamWorkflow(_s(row, "workflow_status", item.workflow_status.value if item.workflow_status else "draft"))
    return action


async def _import_exam_subject(db, institution_id, row):
    year = await _year(db, institution_id, _s(row, "academic_year_label"))
    exam = await _one(db, Exam, Exam.institution_id == institution_id, Exam.academic_year_id == year.id, Exam.name == _s(row, "exam_name"))
    if not exam:
        raise NotFoundError(f"Exam not found: {_s(row, 'exam_name')}")
    subject = await _subject(db, institution_id, _s(row, "subject_code"))
    item = await _one(db, ExamSubject, ExamSubject.exam_id == exam.id, ExamSubject.subject_id == subject.id)
    action = "updated" if item else "created"
    if not item:
        item = ExamSubject(exam_id=exam.id, subject_id=subject.id, max_marks=_i(row, "max_marks", 100), pass_marks=_i(row, "pass_marks", 40))
        db.add(item)
    item.max_marks = _i(row, "max_marks", item.max_marks)
    item.pass_marks = _i(row, "pass_marks", item.pass_marks)
    item.exam_date = _d(row, "exam_date")
    return action


async def _import_mark(db, institution_id, row):
    year = await _year(db, institution_id, _s(row, "academic_year_label"))
    exam = await _one(db, Exam, Exam.institution_id == institution_id, Exam.academic_year_id == year.id, Exam.name == _s(row, "exam_name"))
    if not exam:
        raise NotFoundError(f"Exam not found: {_s(row, 'exam_name')}")
    if exam.workflow_status != ExamWorkflow.DRAFT:
        raise ValidationError("Marks can be imported only while exam is in draft status")
    subject = await _subject(db, institution_id, _s(row, "subject_code"))
    exam_subject = await _one(db, ExamSubject, ExamSubject.exam_id == exam.id, ExamSubject.subject_id == subject.id)
    if not exam_subject:
        raise NotFoundError("Exam subject mapping not found")
    student = await _student(db, _s(row, "roll_number"))
    student_user = await _one(db, User, User.id == student.user_id, User.institution_id == institution_id)
    if not student_user:
        raise ValidationError("Student does not belong to your institution")
    item = await _one(db, Mark, Mark.exam_subject_id == exam_subject.id, Mark.student_id == student.id)
    if item and item.is_locked:
        raise ValidationError("Cannot overwrite a locked mark")
    is_absent = _b(row.get("is_absent"))
    marks = None if is_absent else _f(row, "marks_obtained")
    if not is_absent and (marks is None or marks < 0 or marks > exam_subject.max_marks):
        raise ValidationError(f"Marks must be between 0 and {exam_subject.max_marks}")
    action = "updated" if item else "created"
    if not item:
        item = Mark(exam_subject_id=exam_subject.id, student_id=student.id)
        db.add(item)
    item.marks_obtained = marks
    item.is_absent = is_absent
    item.is_locked = False
    return action


async def _import_library_book(db, institution_id, row):
    isbn = _s(row, "isbn")
    if not isbn:
        raise ValidationError("isbn is required")
    item = await _one(db, Book, Book.isbn == isbn)
    action = "updated" if item else "created"
    total_copies = _i(row, "total_copies", 1)
    available_copies = _i(row, "available_copies", total_copies)
    if not item:
        item = Book(
            institution_id=institution_id,
            isbn=isbn,
            title=_s(row, "title"),
            author=_s(row, "author"),
            publisher=_s(row, "publisher") or None,
            total_copies=total_copies,
            available_copies=available_copies,
        )
        db.add(item)
    item.institution_id = institution_id
    item.title = _s(row, "title", item.title)
    item.author = _s(row, "author", item.author)
    item.publisher = _s(row, "publisher") or item.publisher
    item.total_copies = total_copies
    item.available_copies = min(available_copies, total_copies)
    return action


async def _export_empty(db, institution_id):
    return []


async def _export_library_books(db, institution_id):
    rows = (
        await db.execute(
            select(Book)
            .where(Book.institution_id == institution_id)
            .order_by(Book.title)
        )
    ).scalars().all()
    return [
        {
            "isbn": book.isbn,
            "barcode": book.isbn,
            "title": book.title,
            "author": book.author,
            "publisher": book.publisher or "",
            "total_copies": book.total_copies,
            "available_copies": book.available_copies,
        }
        for book in rows
    ]


async def _export_students(db, institution_id):
    rows = (
        await db.execute(
            select(Student, User, StudentAcademicRecord, Section, Class, Course, Branch, AcademicYear)
            .join(User, User.id == Student.user_id)
            .outerjoin(StudentAcademicRecord, (StudentAcademicRecord.student_id == Student.id) & (StudentAcademicRecord.exited_at == None))
            .outerjoin(Section, Section.id == StudentAcademicRecord.section_id)
            .outerjoin(Class, Class.id == Section.class_id)
            .outerjoin(Course, Course.id == Class.course_id)
            .outerjoin(Branch, Branch.id == StudentAcademicRecord.branch_id)
            .outerjoin(AcademicYear, AcademicYear.id == StudentAcademicRecord.academic_year_id)
            .where(User.institution_id == institution_id)
        )
    ).all()
    return [
        {
            "full_name": user.full_name,
            "email": user.email,
            "username": user.username,
            "password": "",
            "phone": user.phone or "",
            "roll_number": student.roll_number,
            "date_of_birth": student.date_of_birth or "",
            "gender": student.gender or "",
            "guardian_name": student.guardian_name or "",
            "guardian_phone": student.guardian_phone or "",
            "guardian_email": student.guardian_email or "",
            "academic_year_label": year.label if year else "",
            "course_code": course.code if course else "",
            "branch_code": branch.code if branch else "",
            "class_name": class_.name if class_ else "",
            "section_name": section.name if section else "",
        }
        for student, user, _, section, class_, course, branch, year in rows
    ]


async def _export_simple(db, institution_id, model, headers, extra: Callable | None = None):
    rows = (await db.execute(select(model))).scalars().all()
    return [extra(item) if extra else {header: getattr(item, header, "") for header in headers} for item in rows]


SPECS = {
    "academic-years": {"headers": ["label", "start_date", "end_date", "is_current", "is_active"], "examples": [{"label": "2027-28", "start_date": "2027-07-01", "end_date": "2028-06-30", "is_current": "false", "is_active": "true"}]},
    "courses": {"headers": ["code", "name", "level", "duration_years", "is_active"], "examples": [{"code": "BCA", "name": "Bachelor of Computer Applications", "level": "UG", "duration_years": "3", "is_active": "true"}]},
    "branches": {"headers": ["course_code", "code", "name", "is_active"], "examples": [{"course_code": "BTECH", "code": "AIML", "name": "Artificial Intelligence and Machine Learning", "is_active": "true"}]},
    "classes": {"headers": ["course_code", "branch_code", "academic_year_label", "name", "year_no", "semester", "intake_capacity"], "examples": [{"course_code": "BTECH", "branch_code": "CSE", "academic_year_label": "2026-27", "name": "CSE Fourth Year - Semester 7", "year_no": "4", "semester": "7", "intake_capacity": "60"}]},
    "sections": {"headers": ["course_code", "branch_code", "class_name", "name", "max_strength"], "examples": [{"course_code": "BTECH", "branch_code": "CSE", "class_name": "CSE Second Year - Semester 3", "name": "C", "max_strength": "60"}]},
    "subjects": {"headers": ["course_code", "class_name", "branch_code", "academic_year_label", "semester", "code", "name", "credits", "is_active"], "examples": [{"course_code": "BTECH", "class_name": "CSE Second Year - Semester 3", "branch_code": "CSE", "academic_year_label": "2026-27", "semester": "3", "code": "CS2309", "name": "Design and Analysis of Algorithms", "credits": "4", "is_active": "true"}]},
    "students": {"headers": ["full_name", "email", "username", "password", "phone", "roll_number", "date_of_birth", "gender", "guardian_name", "guardian_phone", "guardian_email", "academic_year_label", "course_code", "branch_code", "class_name", "section_name"], "examples": [{"full_name": "New Student", "email": "new.student@student.demo.edu", "username": "new.student", "password": "Demo@123", "phone": "9000099999", "roll_number": "DEC27CSEA099", "date_of_birth": "2007-05-20", "gender": "female", "guardian_name": "Demo Parent", "guardian_phone": "9000099998", "guardian_email": "demo.parent@parent.demo.edu", "academic_year_label": "2026-27", "course_code": "BTECH", "branch_code": "CSE", "class_name": "CSE Second Year - Semester 3", "section_name": "A"}]},
    "teachers": {"headers": ["full_name", "email", "username", "password", "phone", "employee_code", "designation", "joined_at"], "examples": [{"full_name": "Demo Faculty", "email": "demo.faculty@demo.edu", "username": "demo.faculty", "password": "Demo@123", "phone": "9000088888", "employee_code": "FAC-DEMO-01", "designation": "Assistant Professor", "joined_at": "2026-06-01"}]},
    "fee-types": {"headers": ["name", "description"], "examples": [{"name": "Development Fee", "description": "Annual development fee"}]},
    "fee-structures": {"headers": ["fee_type_name", "course_code", "academic_year_label", "amount", "frequency"], "examples": [{"fee_type_name": "Development Fee", "course_code": "BTECH", "academic_year_label": "2026-27", "amount": "12000", "frequency": "annual"}]},
    "student-fees": {"headers": ["roll_number", "fee_type_name", "course_code", "academic_year_label", "amount_due", "due_date"], "examples": [{"roll_number": "DEC26CSEA001", "fee_type_name": "Development Fee", "course_code": "BTECH", "academic_year_label": "2026-27", "amount_due": "12000", "due_date": "2026-08-15"}]},
    "student-documents": {"headers": ["roll_number", "document_type", "title", "file_name", "file_url", "status", "remarks"], "examples": [{"roll_number": "DEC26CSEA001", "document_type": "identity", "title": "Aadhaar", "file_name": "aadhaar.pdf", "file_url": "", "status": "verified", "remarks": "Verified"}]},
    "exams": {"headers": ["academic_year_label", "name", "exam_type", "workflow_status"], "examples": [{"academic_year_label": "2026-27", "name": "Unit Test 1", "exam_type": "unit_test", "workflow_status": "draft"}]},
    "exam-subjects": {"headers": ["exam_name", "academic_year_label", "subject_code", "max_marks", "pass_marks", "exam_date"], "examples": [{"exam_name": "Unit Test 1", "academic_year_label": "2026-27", "subject_code": "CS2301", "max_marks": "50", "pass_marks": "20", "exam_date": "2026-08-20"}]},
    "marks": {"headers": ["exam_name", "academic_year_label", "subject_code", "roll_number", "marks_obtained", "is_absent", "is_locked"], "examples": [{"exam_name": "Unit Test 1", "academic_year_label": "2026-27", "subject_code": "CS2301", "roll_number": "DEC26CSEA001", "marks_obtained": "42", "is_absent": "false", "is_locked": "false"}]},
    "library-books": {"headers": ["isbn", "barcode", "title", "author", "publisher", "total_copies", "available_copies"], "examples": [{"isbn": "9789355420669", "barcode": "9789355420669", "title": "Database System Concepts", "author": "Abraham Silberschatz, Henry F. Korth, S. Sudarshan", "publisher": "McGraw Hill", "total_copies": "6", "available_copies": "6"}]},
}

IMPORTERS = {
    "academic-years": _import_academic_year,
    "courses": _import_course,
    "branches": _import_branch,
    "classes": _import_class,
    "sections": _import_section,
    "subjects": _import_subject,
    "students": _import_student,
    "teachers": _import_teacher,
    "fee-types": _import_fee_type,
    "fee-structures": _import_fee_structure,
    "student-fees": _import_student_fee,
    "student-documents": _import_student_document,
    "exams": _import_exam,
    "exam-subjects": _import_exam_subject,
    "marks": _import_mark,
    "library-books": _import_library_book,
}

EXPORTERS = {key: _export_empty for key in SPECS}
EXPORTERS["students"] = _export_students
EXPORTERS["library-books"] = _export_library_books
