from datetime import date, time
import uuid
from io import BytesIO, StringIO
import csv
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, literal, or_, case
from sqlalchemy.orm import aliased
from app.modules.teachers.model import (
    Teacher,
    TeacherSubject,
    TeacherClass,
    ClassMentor,
    TeacherTimetable,
    TimetableDay,
    HODLink,
    TeacherHODSubjectLink,
)
from app.modules.users.model import User
from app.modules.roles.model import Role, UserRole
from app.modules.academic.model import Class, Branch, Course, AcademicYear, Section, Subject
from app.modules.attendance.model import AttendanceSession, AttendanceRecord, AttendanceStatus
from app.modules.exams.model import Mark, ExamSubject, Exam
from app.modules.students.model import StudentAcademicRecord, StudentStatus
from app.modules.teachers.schema import (
    TeacherCreate,
    SubjectAssignRequest,
    TeacherClassAssignRequest,
    ClassMentorCreate,
    ClassMentorUpdate,
    TeacherTimetableCreate,
    TeacherTimetableUpdate,
)
from app.core.security import hash_password
from app.core.exceptions import NotFoundError, ConflictError, ValidationError, BusinessRuleError, ForbiddenError
from app.core.role_context import has_any_role


async def create_teacher(
    db: AsyncSession, data: TeacherCreate, actor_institution_id: str | None = None
) -> Teacher:
    ex_code = (
        await db.execute(select(Teacher).where(Teacher.employee_code == data.employee_code))
    ).scalar_one_or_none()
    if ex_code:
        raise ConflictError("Employee code already in use")

    if data.user_id:
        user = (
            await db.execute(select(User).where(User.id == data.user_id))
        ).scalar_one_or_none()
        if not user:
            raise NotFoundError("User not found")
        if actor_institution_id and str(user.institution_id) != actor_institution_id:
            raise ValidationError("Selected user does not belong to your institution")

        existing_teacher = (
            await db.execute(select(Teacher).where(Teacher.user_id == user.id))
        ).scalar_one_or_none()
        if existing_teacher:
            raise ConflictError("Teacher profile already exists for this user")

        has_teacher_role = await _user_has_teacher_role(db, str(user.id))
        if not has_teacher_role:
            raise ValidationError("Selected user does not have the Teacher role")
    else:
        institution_id = actor_institution_id or str(data.institution_id)
        ex = (await db.execute(select(User).where(User.email == data.email))).scalar_one_or_none()
        if ex:
            raise ConflictError("Email already registered")
        ex2 = (
            await db.execute(select(User).where(User.username == data.username))
        ).scalar_one_or_none()
        if ex2:
            raise ConflictError("Username already taken")

        user = User(
            institution_id=institution_id,
            email=data.email,
            username=data.username,
            password_hash=hash_password(data.password),
            full_name=data.full_name,
            phone=data.phone,
        )
        db.add(user)
        await db.flush()

        teacher_role = await _get_teacher_role(db, institution_id)
        if teacher_role:
            db.add(UserRole(user_id=user.id, role_id=teacher_role.id))
            await db.flush()

    teacher = Teacher(
        user_id=user.id,
        employee_code=data.employee_code,
        designation=data.designation,
        joined_at=data.joined_at,
    )
    db.add(teacher)
    await db.flush()
    await db.refresh(teacher)
    return teacher


def _has_admin_teacher_scope(current_user: dict | None) -> bool:
    return bool(
        current_user
        and (
            current_user.get("is_superuser")
            or has_any_role(current_user, {"superadmin", "admin", "principal"})
        )
    )


async def _current_teacher_profile(db: AsyncSession, current_user: dict | None) -> Teacher:
    if not current_user:
        raise ForbiddenError("Teacher profile is required")
    teacher = (
        await db.execute(select(Teacher).where(Teacher.user_id == current_user["id"]))
    ).scalar_one_or_none()
    if not teacher:
        raise ForbiddenError("Teacher profile not found for current user")
    return teacher


def _managed_branch_ids_for_hod(teacher_id):
    return select(HODLink.branch_id).where(HODLink.hod_teacher_id == teacher_id)


def _teacher_ids_for_hod_scope(hod_teacher_id):
    managed_branch_ids = _managed_branch_ids_for_hod(hod_teacher_id)
    linked_teacher_ids = (
        select(TeacherHODSubjectLink.teacher_id)
        .join(HODLink, HODLink.id == TeacherHODSubjectLink.hod_link_id)
        .where(HODLink.hod_teacher_id == hod_teacher_id)
    )
    class_teacher_ids = (
        select(TeacherClass.teacher_id)
        .join(Class, Class.id == TeacherClass.class_id)
        .where(Class.branch_id.in_(managed_branch_ids))
    )
    timetable_teacher_ids = (
        select(TeacherTimetable.teacher_id)
        .join(Class, Class.id == TeacherTimetable.class_id)
        .where(Class.branch_id.in_(managed_branch_ids))
    )
    return linked_teacher_ids.union(class_teacher_ids, timetable_teacher_ids)


async def assert_can_view_teacher(db: AsyncSession, current_user: dict | None, teacher_id: str) -> None:
    if _has_admin_teacher_scope(current_user):
        return
    actor_teacher = await _current_teacher_profile(db, current_user)
    if str(actor_teacher.id) == str(teacher_id):
        return
    if has_any_role(current_user or {}, {"hod"}):
        allowed_ids = _teacher_ids_for_hod_scope(actor_teacher.id)
        row = (
            await db.execute(
                select(Teacher.id)
                .where(Teacher.id == teacher_id, Teacher.id.in_(allowed_ids))
                .limit(1)
            )
        ).first()
        if row:
            return
    raise ForbiddenError("You can view only linked teacher data")


async def list_teachers(
    db: AsyncSession,
    institution_id: str,
    offset: int,
    limit: int,
    current_user: dict | None = None,
):
    q = (
        select(Teacher)
        .join(User, Teacher.user_id == User.id)
        .where(User.institution_id == institution_id)
    )
    if current_user and not _has_admin_teacher_scope(current_user):
        actor_teacher = await _current_teacher_profile(db, current_user)
        if has_any_role(current_user, {"hod"}):
            q = q.where(
                or_(
                    Teacher.id == actor_teacher.id,
                    Teacher.id.in_(_teacher_ids_for_hod_scope(actor_teacher.id)),
                )
            )
        else:
            q = q.where(Teacher.id == actor_teacher.id)
    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar()
    result = await db.execute(q.order_by(Teacher.employee_code.asc()).offset(offset).limit(limit))
    return result.scalars().all(), total


async def get_teacher(db: AsyncSession, teacher_id: str) -> Teacher:
    obj = (await db.execute(select(Teacher).where(Teacher.id == teacher_id))).scalar_one_or_none()
    if not obj:
        raise NotFoundError("Teacher not found")
    return obj


async def list_teacher_candidates(db: AsyncSession, institution_id: str):
    teacher_user_ids = (
        select(Teacher.user_id)
        .subquery()
    )
    rows = (
        await db.execute(
            select(User)
            .join(UserRole, UserRole.user_id == User.id)
            .join(Role, Role.id == UserRole.role_id)
            .where(
                User.institution_id == institution_id,
                _teacher_role_filter(),
                ~User.id.in_(select(teacher_user_ids.c.user_id)),
            )
            .order_by(User.full_name.asc())
        )
    ).scalars().all()
    return rows


async def list_hod_teacher_candidates(db: AsyncSession, institution_id: str):
    rows = (
        await db.execute(
            select(User.id, Teacher.id, User.full_name, Teacher.employee_code)
            .select_from(User)
            .outerjoin(Teacher, Teacher.user_id == User.id)
            .join(UserRole, UserRole.user_id == User.id)
            .join(Role, Role.id == UserRole.role_id)
            .where(
                User.institution_id == institution_id,
                or_(
                    func.lower(Role.slug) == "hod",
                    func.lower(Role.name) == "hod",
                    func.lower(Role.name).like("%hod%"),
                ),
            )
            .order_by(User.full_name.asc())
        )
    ).all()
    return [
        {
            "user_id": str(row[0]),
            "teacher_id": str(row[1]) if row[1] else None,
            "full_name": row[2],
            "employee_code": row[3] or "-",
        }
        for row in rows
    ]


async def assign_subject(db: AsyncSession, teacher_id: str, data: SubjectAssignRequest) -> TeacherSubject:
    await get_teacher(db, teacher_id)
    ts = TeacherSubject(
        teacher_id=teacher_id,
        subject_id=data.subject_id,
        section_id=data.section_id,
        academic_year_id=data.academic_year_id,
    )
    db.add(ts)
    await db.flush()
    await db.refresh(ts)
    return ts


async def list_teacher_classes(db: AsyncSession, teacher_id: str) -> list[dict]:
    await get_teacher(db, teacher_id)

    rows = (
        await db.execute(
            select(
                TeacherClass.id,
                Class.id.label("class_id"),
                Class.name.label("class_name"),
                Class.semester,
                Branch.id.label("branch_id"),
                Branch.name.label("branch_name"),
                Class.academic_year_id,
                AcademicYear.label.label("academic_year_label"),
            )
            .join(Class, Class.id == TeacherClass.class_id)
            .join(Branch, Branch.id == Class.branch_id)
            .outerjoin(AcademicYear, AcademicYear.id == Class.academic_year_id)
            .where(TeacherClass.teacher_id == teacher_id)
            .order_by(Branch.name.asc(), Class.semester.asc(), Class.name.asc())
        )
    ).mappings().all()
    return [dict(row) for row in rows]


async def _class_section_context(db: AsyncSession, class_id: str, section_id: str, academic_year_id: str):
    class_row = (
        await db.execute(
            select(Class, Section, Branch, Course)
            .join(Section, Section.class_id == Class.id)
            .outerjoin(Branch, Branch.id == Class.branch_id)
            .join(Course, Course.id == Class.course_id)
            .where(
                Class.id == class_id,
                Section.id == section_id,
                Section.class_id == Class.id,
            )
        )
    ).first()
    year = (
        await db.execute(select(AcademicYear).where(AcademicYear.id == academic_year_id))
    ).scalar_one_or_none()
    if not class_row or not year:
        raise ValidationError("Selected academic year/class/section is invalid")
    class_, section, branch, course = class_row
    return class_, section, branch, course, year


async def _assert_hod_can_manage_class(
    db: AsyncSession,
    current_user: dict,
    class_id: str,
    section_id: str,
    academic_year_id: str,
):
    class_, section, branch, course, year = await _class_section_context(db, class_id, section_id, academic_year_id)
    if current_user.get("is_superuser") or has_any_role(current_user, {"admin", "superadmin", "principal"}):
        return class_, section, branch, course, year

    actor_teacher = (
        await db.execute(select(Teacher).where(Teacher.user_id == current_user["id"]))
    ).scalar_one_or_none()
    if not actor_teacher:
        raise ForbiddenError("Only Admin/HOD can manage class mentors")

    hod_link = (
        await db.execute(
            select(HODLink.id).where(
                HODLink.hod_teacher_id == actor_teacher.id,
                HODLink.institution_id == current_user["institution_id"],
                HODLink.course_id == class_.course_id,
                HODLink.branch_id == class_.branch_id,
            )
        )
    ).scalar_one_or_none()
    if not hod_link:
        raise ForbiddenError("Only the mapped HOD can manage mentors for this class/section")
    return class_, section, branch, course, year


async def create_class_mentor(db: AsyncSession, payload: ClassMentorCreate, current_user: dict) -> ClassMentor:
    await get_teacher(db, str(payload.teacher_id))
    await _assert_hod_can_manage_class(db, current_user, str(payload.class_id), str(payload.section_id), str(payload.academic_year_id))
    existing = (
        await db.execute(
            select(ClassMentor).where(
                ClassMentor.academic_year_id == payload.academic_year_id,
                ClassMentor.class_id == payload.class_id,
                ClassMentor.section_id == payload.section_id,
            )
        )
    ).scalar_one_or_none()
    if existing:
        raise ConflictError("Mentor already assigned for selected academic year/class/section")
    mentor = ClassMentor(
        teacher_id=payload.teacher_id,
        academic_year_id=payload.academic_year_id,
        class_id=payload.class_id,
        section_id=payload.section_id,
        assigned_by_user_id=current_user["id"],
        is_active=True,
    )
    db.add(mentor)
    await db.flush()
    await db.refresh(mentor)
    return mentor


async def update_class_mentor(db: AsyncSession, mentor_id: str, payload: ClassMentorUpdate, current_user: dict) -> ClassMentor:
    mentor = (await db.execute(select(ClassMentor).where(ClassMentor.id == mentor_id))).scalar_one_or_none()
    if not mentor:
        raise NotFoundError("Class mentor assignment not found")
    teacher_id = payload.teacher_id or mentor.teacher_id
    academic_year_id = payload.academic_year_id or mentor.academic_year_id
    class_id = payload.class_id or mentor.class_id
    section_id = payload.section_id or mentor.section_id
    await get_teacher(db, str(teacher_id))
    await _assert_hod_can_manage_class(db, current_user, str(class_id), str(section_id), str(academic_year_id))
    duplicate = (
        await db.execute(
            select(ClassMentor.id).where(
                ClassMentor.id != mentor.id,
                ClassMentor.academic_year_id == academic_year_id,
                ClassMentor.class_id == class_id,
                ClassMentor.section_id == section_id,
            )
        )
    ).scalar_one_or_none()
    if duplicate:
        raise ConflictError("Mentor already assigned for selected academic year/class/section")
    mentor.teacher_id = teacher_id
    mentor.academic_year_id = academic_year_id
    mentor.class_id = class_id
    mentor.section_id = section_id
    if payload.is_active is not None:
        mentor.is_active = payload.is_active
    await db.flush()
    await db.refresh(mentor)
    return mentor


async def delete_class_mentor(db: AsyncSession, mentor_id: str, current_user: dict) -> None:
    mentor = (await db.execute(select(ClassMentor).where(ClassMentor.id == mentor_id))).scalar_one_or_none()
    if not mentor:
        raise NotFoundError("Class mentor assignment not found")
    await _assert_hod_can_manage_class(db, current_user, str(mentor.class_id), str(mentor.section_id), str(mentor.academic_year_id))
    await db.delete(mentor)
    await db.flush()


async def list_class_mentors(
    db: AsyncSession,
    current_user: dict,
    academic_year_id: str | None = None,
    branch_id: str | None = None,
    class_id: str | None = None,
    section_id: str | None = None,
) -> list[dict]:
    q = (
        select(
            ClassMentor.id,
            ClassMentor.teacher_id,
            User.full_name.label("teacher_name"),
            Teacher.employee_code,
            ClassMentor.academic_year_id,
            AcademicYear.label.label("academic_year_label"),
            Course.id.label("course_id"),
            Course.name.label("course_name"),
            Branch.id.label("branch_id"),
            Branch.name.label("branch_name"),
            ClassMentor.class_id,
            Class.name.label("class_name"),
            ClassMentor.section_id,
            Section.name.label("section_name"),
            ClassMentor.assigned_by_user_id,
            ClassMentor.is_active,
            ClassMentor.created_at,
        )
        .select_from(ClassMentor)
        .join(Teacher, Teacher.id == ClassMentor.teacher_id)
        .join(User, User.id == Teacher.user_id)
        .join(AcademicYear, AcademicYear.id == ClassMentor.academic_year_id)
        .join(Class, Class.id == ClassMentor.class_id)
        .join(Section, Section.id == ClassMentor.section_id)
        .outerjoin(Branch, Branch.id == Class.branch_id)
        .join(Course, Course.id == Class.course_id)
        .where(User.institution_id == current_user["institution_id"])
    )
    if academic_year_id:
        q = q.where(ClassMentor.academic_year_id == academic_year_id)
    if branch_id:
        q = q.where(Class.branch_id == branch_id)
    if class_id:
        q = q.where(ClassMentor.class_id == class_id)
    if section_id:
        q = q.where(ClassMentor.section_id == section_id)
    if not (current_user.get("is_superuser") or has_any_role(current_user, {"admin", "superadmin", "principal"})):
        actor_teacher = (await db.execute(select(Teacher).where(Teacher.user_id == current_user["id"]))).scalar_one_or_none()
        if not actor_teacher:
            raise ForbiddenError("Only Admin/HOD can view class mentors")
        hod_branches = select(HODLink.branch_id).where(HODLink.hod_teacher_id == actor_teacher.id)
        q = q.where(Class.branch_id.in_(hod_branches))
    rows = (await db.execute(q.order_by(AcademicYear.label.desc(), Branch.name.asc(), Class.name.asc(), Section.name.asc()))).mappings().all()
    return [dict(row) for row in rows]


async def list_teacher_teaching_scope(
    db: AsyncSession,
    teacher_id: str,
    branch_id: str | None = None,
    class_id: str | None = None,
    section_id: str | None = None,
) -> dict:
    await get_teacher(db, teacher_id)

    teacher_class_ids = select(TeacherClass.class_id).where(TeacherClass.teacher_id == teacher_id)
    teacher_class_branch_ids = (
        select(Class.branch_id)
        .join(TeacherClass, TeacherClass.class_id == Class.id)
        .where(TeacherClass.teacher_id == teacher_id)
    )
    linked_branch_ids = (
        select(HODLink.branch_id)
        .join(TeacherHODSubjectLink, TeacherHODSubjectLink.hod_link_id == HODLink.id)
        .where(TeacherHODSubjectLink.teacher_id == teacher_id)
    )

    branch_q = (
        select(Branch.id, Branch.name)
        .join(Course, Course.id == Branch.course_id)
        .add_columns(Course.id.label("course_id"), Course.name.label("course_name"))
        .where(
            or_(
                Branch.id.in_(teacher_class_branch_ids),
                Branch.id.in_(linked_branch_ids),
            )
        )
        .distinct()
        .order_by(Branch.name.asc())
    )
    branches = (await db.execute(branch_q)).mappings().all()

    class_q = (
        select(
            Class.id,
            Class.name,
            Class.semester,
            Class.year_no,
            Class.branch_id,
            Branch.name.label("branch_name"),
            Class.course_id,
            Course.name.label("course_name"),
        )
        .outerjoin(Branch, Branch.id == Class.branch_id)
        .outerjoin(Course, Course.id == Class.course_id)
        .where(
            or_(
                Class.id.in_(teacher_class_ids),
                Class.branch_id.in_(linked_branch_ids),
            )
        )
    )
    if branch_id:
        class_q = class_q.where(Class.branch_id == branch_id)
    classes = (
        await db.execute(class_q.order_by(Branch.name.asc(), Class.semester.asc(), Class.name.asc()))
    ).mappings().all()

    section_q = (
        select(
            Section.id,
            Section.name,
            Section.class_id,
            Class.name.label("class_name"),
        )
        .join(Class, Class.id == Section.class_id)
        .outerjoin(TeacherSubject, TeacherSubject.section_id == Section.id)
        .where(
            or_(
                TeacherSubject.teacher_id == teacher_id,
                Class.id.in_(teacher_class_ids),
                Class.branch_id.in_(linked_branch_ids),
            )
        )
        .distinct()
    )
    if branch_id:
        section_q = section_q.where(Class.branch_id == branch_id)
    if class_id:
        section_q = section_q.where(Section.class_id == class_id)
    sections = (
        await db.execute(section_q.order_by(Class.name.asc(), Section.name.asc()))
    ).mappings().all()

    subject_q = (
        select(
            Subject.id,
            Subject.name,
            Subject.code,
            Subject.class_id,
            func.coalesce(Subject.branch_id, Class.branch_id).label("branch_id"),
            TeacherSubject.section_id,
        )
        .outerjoin(TeacherSubject, TeacherSubject.subject_id == Subject.id)
        .outerjoin(TeacherHODSubjectLink, TeacherHODSubjectLink.subject_id == Subject.id)
        .outerjoin(HODLink, HODLink.id == TeacherHODSubjectLink.hod_link_id)
        .join(Class, Class.id == Subject.class_id)
        .where(
            or_(
                TeacherSubject.teacher_id == teacher_id,
                TeacherHODSubjectLink.teacher_id == teacher_id,
            )
        )
        .distinct()
    )
    if branch_id:
        subject_q = subject_q.where(
            or_(
                Class.branch_id == branch_id,
                Subject.branch_id == branch_id,
                HODLink.branch_id == branch_id,
            )
        )
    if class_id:
        subject_q = subject_q.where(Subject.class_id == class_id)
    if section_id:
        section_class_id = (
            await db.execute(select(Section.class_id).where(Section.id == section_id))
        ).scalar_one_or_none()
        subject_q = subject_q.where(
            or_(
                TeacherSubject.section_id == section_id,
                Subject.class_id == section_class_id,
            )
        )
    subjects = (
        await db.execute(subject_q.order_by(Subject.name.asc(), Subject.code.asc()))
    ).mappings().all()

    return {
        "branches": [dict(row) for row in branches],
        "classes": [dict(row) for row in classes],
        "sections": [dict(row) for row in sections],
        "subjects": [dict(row) for row in subjects],
    }


async def list_teacher_timetable(
    db: AsyncSession,
    teacher_id: str,
    day_of_week: TimetableDay | None = None,
    session_date: date | None = None,
) -> list[dict]:
    await get_teacher(db, teacher_id)

    q = (
        select(
            TeacherTimetable.id,
            TeacherTimetable.teacher_id,
            TeacherTimetable.class_id,
            Class.name.label("class_name"),
            TeacherTimetable.section_id,
            Section.name.label("section_name"),
            TeacherTimetable.subject_id,
            Subject.name.label("subject_name"),
            Branch.id.label("branch_id"),
            Branch.name.label("branch_name"),
            TeacherTimetable.academic_year_id,
            AcademicYear.label.label("academic_year_label"),
            TeacherTimetable.day_of_week,
            TeacherTimetable.start_time,
            TeacherTimetable.end_time,
            TeacherTimetable.room_no,
            TeacherTimetable.version_no,
            TeacherTimetable.is_active,
        )
        .join(Class, Class.id == TeacherTimetable.class_id)
        .join(Section, Section.id == TeacherTimetable.section_id)
        .join(Subject, Subject.id == TeacherTimetable.subject_id)
        .join(Branch, Branch.id == Class.branch_id)
        .join(AcademicYear, AcademicYear.id == TeacherTimetable.academic_year_id)
        .where(TeacherTimetable.teacher_id == teacher_id)
    )
    q = q.where(TeacherTimetable.is_active == True)

    if session_date:
        q = q.add_columns(
            AttendanceSession.id.label("session_id"),
            AttendanceSession.status.label("session_status"),
            AttendanceSession.session_date,
        ).outerjoin(
            AttendanceSession,
            (AttendanceSession.timetable_id == TeacherTimetable.id)
            & (AttendanceSession.session_date == session_date),
        )
    else:
        q = q.add_columns(
            literal(None).label("session_id"),
            literal(None).label("session_status"),
            literal(None).label("session_date"),
        )

    target_day = day_of_week or (_date_to_day(session_date) if session_date else None)
    if target_day:
        q = q.where(TeacherTimetable.day_of_week == target_day)

    rows = (
        await db.execute(
            q.order_by(
                TeacherTimetable.day_of_week.asc(),
                TeacherTimetable.start_time.asc(),
                Class.name.asc(),
                Section.name.asc(),
            )
        )
    ).mappings().all()
    return [dict(row) for row in rows]


async def assign_class(
    db: AsyncSession, teacher_id: str, data: TeacherClassAssignRequest
) -> TeacherClass:
    teacher = await get_teacher(db, teacher_id)
    class_obj = (
        await db.execute(
            select(Class, Course.institution_id)
            .join(Branch, Branch.id == Class.branch_id)
            .join(Course, Course.id == Branch.course_id)
            .where(Class.id == data.class_id)
        )
    ).first()
    if not class_obj:
        raise NotFoundError("Class not found")

    class_row, class_institution_id = class_obj
    user = (
        await db.execute(select(User).where(User.id == teacher.user_id))
    ).scalar_one()
    if class_institution_id != user.institution_id:
        raise ValidationError("Class does not belong to the teacher's institution")

    existing = (
        await db.execute(
            select(TeacherClass).where(
                TeacherClass.teacher_id == teacher.id,
                TeacherClass.class_id == data.class_id,
            )
        )
    ).scalar_one_or_none()
    if existing:
        raise ConflictError("Class already assigned to this teacher")

    teacher_class = TeacherClass(teacher_id=teacher.id, class_id=class_row.id)
    db.add(teacher_class)
    await db.flush()
    await db.refresh(teacher_class)
    return teacher_class


async def create_timetable_entry(
    db: AsyncSession, teacher_id: str, data: TeacherTimetableCreate
) -> TeacherTimetable:
    teacher = await get_teacher(db, teacher_id)
    class_obj, section_obj, subject_obj = await _validate_timetable_refs(
        db,
        teacher,
        data.class_id,
        data.section_id,
        data.subject_id,
        data.academic_year_id,
    )

    _validate_time_range(data.start_time, data.end_time)
    await _ensure_timetable_slot_available(
        db,
        teacher.id,
        data.section_id,
        data.day_of_week,
        data.start_time,
        data.end_time,
        exclude_id=None,
    )
    await _ensure_teacher_subject_link(
        db,
        teacher.id,
        data.subject_id,
        data.section_id,
        data.academic_year_id,
    )

    entry = TeacherTimetable(
        teacher_id=teacher.id,
        class_id=class_obj.id,
        section_id=section_obj.id,
        subject_id=subject_obj.id,
        academic_year_id=data.academic_year_id,
        day_of_week=data.day_of_week,
        start_time=data.start_time,
        end_time=data.end_time,
        room_no=data.room_no,
        version_no=data.version_no,
        is_active=data.is_active,
    )
    db.add(entry)
    await db.flush()
    await db.refresh(entry)
    return entry


async def update_timetable_entry(
    db: AsyncSession, entry_id: str, data: TeacherTimetableUpdate
) -> TeacherTimetable:
    entry = await _get_timetable_entry(db, entry_id)
    teacher = await get_teacher(db, str(entry.teacher_id))
    incoming = data.model_dump(exclude_unset=True)

    class_id = incoming.get("class_id", entry.class_id)
    section_id = incoming.get("section_id", entry.section_id)
    subject_id = incoming.get("subject_id", entry.subject_id)
    academic_year_id = incoming.get("academic_year_id", entry.academic_year_id)
    day_of_week = incoming.get("day_of_week", entry.day_of_week)
    start_time = incoming.get("start_time", entry.start_time)
    end_time = incoming.get("end_time", entry.end_time)

    await _validate_timetable_refs(
        db,
        teacher,
        class_id,
        section_id,
        subject_id,
        academic_year_id,
    )
    _validate_time_range(start_time, end_time)
    await _ensure_timetable_slot_available(
        db,
        teacher.id,
        section_id,
        day_of_week,
        start_time,
        end_time,
        exclude_id=entry.id,
    )
    await _ensure_teacher_subject_link(
        db,
        teacher.id,
        subject_id,
        section_id,
        academic_year_id,
    )

    for key, value in incoming.items():
        setattr(entry, key, value)

    await db.flush()
    await db.refresh(entry)
    return entry


async def reassign_timetable_entry(
    db: AsyncSession,
    entry_id: str,
    target_teacher_id: str,
    actor_user_id: str,
    actor_institution_id: str,
    is_superuser: bool = False,
) -> TeacherTimetable:
    entry = await _get_timetable_entry(db, entry_id)
    target_teacher = await get_teacher(db, target_teacher_id)

    row = (
        await db.execute(
            select(Course.institution_id, Branch.id.label("branch_id"))
            .select_from(TeacherTimetable)
            .join(Class, Class.id == TeacherTimetable.class_id)
            .join(Branch, Branch.id == Class.branch_id)
            .join(Course, Course.id == Branch.course_id)
            .where(TeacherTimetable.id == entry.id)
        )
    ).mappings().first()
    if not row:
        raise NotFoundError("Timetable entry not found")

    if str(row["institution_id"]) != actor_institution_id:
        raise ForbiddenError("You cannot reallocate classes outside your institution")

    target_user = (await db.execute(select(User).where(User.id == target_teacher.user_id))).scalar_one()
    if str(target_user.institution_id) != actor_institution_id:
        raise ValidationError("Target teacher belongs to another institution")

    if not is_superuser:
        actor_teacher = (await db.execute(select(Teacher).where(Teacher.user_id == actor_user_id))).scalar_one_or_none()
        if not actor_teacher:
            raise ForbiddenError("Only HOD or super admin can reallocate classes")
        can_manage_branch = (
            await db.execute(
                select(HODLink.id).where(
                    HODLink.hod_teacher_id == actor_teacher.id,
                    HODLink.institution_id == uuid.UUID(actor_institution_id),
                    HODLink.branch_id == row["branch_id"],
                )
            )
        ).first()
        if not can_manage_branch:
            raise ForbiddenError("Only the mapped HOD can reallocate this class")

    await _ensure_timetable_slot_available(
        db,
        target_teacher.id,
        entry.section_id,
        entry.day_of_week,
        entry.start_time,
        entry.end_time,
        exclude_id=entry.id,
    )

    entry.teacher_id = target_teacher.id
    entry.version_no = (entry.version_no or 1) + 1
    await db.flush()
    await db.refresh(entry)
    return entry


async def delete_timetable_entry(db: AsyncSession, entry_id: str) -> None:
    entry = await _get_timetable_entry(db, entry_id)
    active_session = (
        await db.execute(
            select(AttendanceSession.id).where(
                AttendanceSession.teacher_id == entry.teacher_id,
                AttendanceSession.section_id == entry.section_id,
                AttendanceSession.subject_id == entry.subject_id,
                AttendanceSession.academic_year_id == entry.academic_year_id,
            )
        )
    ).first()
    if active_session:
        raise BusinessRuleError("Cannot delete timetable entry with linked attendance sessions")
    await db.delete(entry)


async def import_timetable_entries(
    db: AsyncSession,
    teacher_id: str,
    filename: str,
    content: bytes,
) -> dict:
    teacher = await get_teacher(db, teacher_id)
    rows = _read_timetable_rows(filename, content)
    created = 0
    skipped = 0

    for row in rows:
        try:
            data = TeacherTimetableCreate(
                class_id=row["class_id"],
                section_id=row["section_id"],
                subject_id=row["subject_id"],
                academic_year_id=row["academic_year_id"],
                day_of_week=row["day_of_week"],
                start_time=_parse_time(row["start_time"]),
                end_time=_parse_time(row["end_time"]),
                room_no=row.get("room_no") or None,
                version_no=int(row.get("version_no") or 1),
                is_active=str(row.get("is_active", "true")).lower() in ("true", "1", "yes"),
            )
            await create_timetable_entry(db, str(teacher.id), data)
            created += 1
        except Exception:
            skipped += 1

    return {"created": created, "skipped": skipped, "total": len(rows)}


async def remove_class(db: AsyncSession, teacher_id: str, class_id: str) -> None:
    teacher = await get_teacher(db, teacher_id)
    teacher_class = (
        await db.execute(
            select(TeacherClass).where(
                TeacherClass.teacher_id == teacher.id,
                TeacherClass.class_id == class_id,
            )
        )
    ).scalar_one_or_none()
    if not teacher_class:
        raise NotFoundError("Teacher class link not found")

    timetable_exists = (
        await db.execute(
            select(TeacherTimetable.id).where(
                TeacherTimetable.teacher_id == teacher.id,
                TeacherTimetable.class_id == class_id,
            )
        )
    ).first()
    if timetable_exists:
        raise BusinessRuleError("Remove timetable entries for this class before unlinking it")

    await db.delete(teacher_class)


async def create_hod_link(
    db: AsyncSession,
    hod_teacher_id,
    institution_id,
    course_id,
    branch_id,
    hod_user_id=None,
) -> HODLink:
    hod_teacher = None
    hod_user = None
    if hod_teacher_id:
        hod_teacher = await get_teacher(db, str(hod_teacher_id))
        hod_user = (await db.execute(select(User).where(User.id == hod_teacher.user_id))).scalar_one()
    elif hod_user_id:
        hod_user = (await db.execute(select(User).where(User.id == hod_user_id))).scalar_one_or_none()
        if not hod_user:
            raise NotFoundError("HOD user not found")
        hod_teacher = (await db.execute(select(Teacher).where(Teacher.user_id == hod_user.id))).scalar_one_or_none()
        if not hod_teacher:
            base = f"HOD-{str(hod_user.id).replace('-', '')[:8]}".upper()
            employee_code = base
            while True:
                exists = (
                    await db.execute(select(Teacher.id).where(Teacher.employee_code == employee_code))
                ).first()
                if not exists:
                    break
                employee_code = f"{base}-{uuid.uuid4().hex[:4].upper()}"
            hod_teacher = Teacher(user_id=hod_user.id, employee_code=employee_code, designation="HOD")
            db.add(hod_teacher)
            await db.flush()
    else:
        raise ValidationError("Provide hod_teacher_id or hod_user_id")

    course = (
        await db.execute(select(Course).where(Course.id == course_id))
    ).scalar_one_or_none()
    if not course:
        raise NotFoundError("Course not found")
    if course.institution_id != institution_id:
        raise ValidationError("Course does not belong to selected institution")

    branch = (
        await db.execute(select(Branch).where(Branch.id == branch_id))
    ).scalar_one_or_none()
    if not branch:
        raise NotFoundError("Branch not found")
    if branch.course_id != course.id:
        raise ValidationError("Branch does not belong to selected course")
    if hod_user.institution_id != institution_id:
        raise ValidationError("Selected HOD teacher does not belong to selected institution")

    existing = (
        await db.execute(
            select(HODLink).where(
                HODLink.hod_teacher_id == hod_teacher.id,
                HODLink.institution_id == institution_id,
                HODLink.course_id == course.id,
                HODLink.branch_id == branch.id,
            )
        )
    ).scalar_one_or_none()
    if existing:
        raise ConflictError("HOD link already exists")

    link = HODLink(
        hod_teacher_id=hod_teacher.id,
        institution_id=institution_id,
        course_id=course.id,
        branch_id=branch.id,
    )
    db.add(link)
    await db.flush()
    await db.refresh(link)
    return link


async def list_hod_links(
    db: AsyncSession,
    institution_id: str | None = None,
    course_id: str | None = None,
    branch_id: str | None = None,
    current_user: dict | None = None,
) -> list[dict]:
    q = (
        select(
            HODLink.id,
            HODLink.hod_teacher_id,
            User.full_name.label("hod_teacher_name"),
            HODLink.institution_id,
            HODLink.course_id,
            Course.name.label("course_name"),
            HODLink.branch_id,
            Branch.name.label("branch_name"),
        )
        .join(Teacher, Teacher.id == HODLink.hod_teacher_id)
        .join(User, User.id == Teacher.user_id)
        .join(Course, Course.id == HODLink.course_id)
        .join(Branch, Branch.id == HODLink.branch_id)
    )
    if institution_id:
        q = q.where(HODLink.institution_id == institution_id)
    if course_id:
        q = q.where(HODLink.course_id == course_id)
    if branch_id:
        q = q.where(HODLink.branch_id == branch_id)
    if current_user and not _has_admin_teacher_scope(current_user):
        actor_teacher = await _current_teacher_profile(db, current_user)
        if has_any_role(current_user, {"hod"}):
            q = q.where(HODLink.hod_teacher_id == actor_teacher.id)
        else:
            visible_hod_link_ids = select(TeacherHODSubjectLink.hod_link_id).where(
                TeacherHODSubjectLink.teacher_id == actor_teacher.id
            )
            q = q.where(HODLink.id.in_(visible_hod_link_ids))

    rows = (await db.execute(q.order_by(User.full_name.asc()))).mappings().all()
    return [dict(row) for row in rows]


async def remove_hod_link(db: AsyncSession, link_id: str) -> None:
    link = (await db.execute(select(HODLink).where(HODLink.id == link_id))).scalar_one_or_none()
    if not link:
        raise NotFoundError("HOD link not found")
    await db.delete(link)


async def update_hod_link(
    db: AsyncSession,
    link_id: str,
    hod_teacher_id,
    institution_id,
    course_id,
    branch_id,
    hod_user_id=None,
) -> HODLink:
    link = (await db.execute(select(HODLink).where(HODLink.id == link_id))).scalar_one_or_none()
    if not link:
        raise NotFoundError("HOD link not found")

    hod_teacher = None
    hod_user = None
    if hod_teacher_id:
        hod_teacher = await get_teacher(db, str(hod_teacher_id))
        hod_user = (await db.execute(select(User).where(User.id == hod_teacher.user_id))).scalar_one()
    elif hod_user_id:
        hod_user = (await db.execute(select(User).where(User.id == hod_user_id))).scalar_one_or_none()
        if not hod_user:
            raise NotFoundError("HOD user not found")
        hod_teacher = (await db.execute(select(Teacher).where(Teacher.user_id == hod_user.id))).scalar_one_or_none()
        if not hod_teacher:
            base = f"HOD-{str(hod_user.id).replace('-', '')[:8]}".upper()
            employee_code = base
            while True:
                exists = (
                    await db.execute(select(Teacher.id).where(Teacher.employee_code == employee_code))
                ).first()
                if not exists:
                    break
                employee_code = f"{base}-{uuid.uuid4().hex[:4].upper()}"
            hod_teacher = Teacher(user_id=hod_user.id, employee_code=employee_code, designation="HOD")
            db.add(hod_teacher)
            await db.flush()
    else:
        raise ValidationError("Provide hod_teacher_id or hod_user_id")

    course = (await db.execute(select(Course).where(Course.id == course_id))).scalar_one_or_none()
    if not course:
        raise NotFoundError("Course not found")
    if course.institution_id != institution_id:
        raise ValidationError("Course does not belong to selected institution")

    branch = (await db.execute(select(Branch).where(Branch.id == branch_id))).scalar_one_or_none()
    if not branch:
        raise NotFoundError("Branch not found")
    if branch.course_id != course.id:
        raise ValidationError("Branch does not belong to selected course")
    if hod_user.institution_id != institution_id:
        raise ValidationError("Selected HOD teacher does not belong to selected institution")

    existing = (
        await db.execute(
            select(HODLink).where(
                HODLink.hod_teacher_id == hod_teacher.id,
                HODLink.institution_id == institution_id,
                HODLink.course_id == course.id,
                HODLink.branch_id == branch.id,
                HODLink.id != link.id,
            )
        )
    ).scalar_one_or_none()
    if existing:
        raise ConflictError("HOD link already exists")

    link.hod_teacher_id = hod_teacher.id
    link.institution_id = institution_id
    link.course_id = course.id
    link.branch_id = branch.id
    await db.flush()
    await db.refresh(link)
    return link


async def create_teacher_hod_subject_links(
    db: AsyncSession, teacher_id, hod_link_id, section_id, subject_ids: list
) -> dict:
    teacher = await get_teacher(db, str(teacher_id))
    teacher_user = (await db.execute(select(User).where(User.id == teacher.user_id))).scalar_one()
    hod_link = (await db.execute(select(HODLink).where(HODLink.id == hod_link_id))).scalar_one_or_none()
    if not hod_link:
        raise NotFoundError("HOD link not found")
    if teacher_user.institution_id != hod_link.institution_id:
        raise ValidationError("Teacher does not belong to selected institution")
    section = (
        await db.execute(
            select(Section)
            .join(Class, Class.id == Section.class_id)
            .where(
                Section.id == section_id,
                Class.course_id == hod_link.course_id,
                Class.branch_id == hod_link.branch_id,
            )
        )
    ).scalar_one_or_none()
    if not section:
        raise ValidationError("Section is required and must belong to selected branch/class")

    created: list[TeacherHODSubjectLink] = []
    skipped: list[str] = []
    for subject_id in subject_ids:
        subject = (await db.execute(select(Subject).where(Subject.id == subject_id))).scalar_one_or_none()
        if not subject:
            raise NotFoundError("Subject not found")
        if subject.course_id != hod_link.course_id:
            raise ValidationError("Subject does not belong to selected course")
        if subject.branch_id and subject.branch_id != hod_link.branch_id:
            raise ValidationError("Subject does not belong to selected branch")
        if subject.class_id and subject.class_id != section.class_id:
            raise ValidationError("Subject does not belong to selected class")

        existing = (
            await db.execute(
                select(TeacherHODSubjectLink).where(
                    TeacherHODSubjectLink.teacher_id == teacher.id,
                    TeacherHODSubjectLink.hod_link_id == hod_link.id,
                    TeacherHODSubjectLink.section_id == section.id,
                    TeacherHODSubjectLink.subject_id == subject.id,
                )
            )
        ).scalar_one_or_none()
        if existing:
            skipped.append(str(subject.id))
            continue

        link = TeacherHODSubjectLink(
            teacher_id=teacher.id,
            hod_link_id=hod_link.id,
            section_id=section.id,
            subject_id=subject.id,
        )
        db.add(link)
        created.append(link)

    await db.flush()
    for item in created:
        await db.refresh(item)
    return {"created": created, "skipped_subject_ids": skipped}


async def list_teacher_hod_subject_links(
    db: AsyncSession,
    institution_id: str | None = None,
    course_id: str | None = None,
    branch_id: str | None = None,
    current_user: dict | None = None,
) -> list[dict]:
    teacher_user = aliased(User)
    hod_teacher = aliased(Teacher)
    hod_user = aliased(User)

    q = (
        select(
            TeacherHODSubjectLink.id,
            TeacherHODSubjectLink.teacher_id,
            teacher_user.full_name.label("teacher_name"),
            HODLink.id.label("hod_link_id"),
            HODLink.hod_teacher_id,
            hod_user.full_name.label("hod_teacher_name"),
            HODLink.institution_id,
            HODLink.course_id,
            Course.name.label("course_name"),
            HODLink.branch_id,
            Branch.name.label("branch_name"),
            Class.id.label("class_id"),
            Class.name.label("class_name"),
            TeacherHODSubjectLink.section_id,
            Section.name.label("section_name"),
            TeacherHODSubjectLink.subject_id,
            Subject.name.label("subject_name"),
        )
        .join(Teacher, Teacher.id == TeacherHODSubjectLink.teacher_id)
        .join(teacher_user, teacher_user.id == Teacher.user_id)
        .join(HODLink, HODLink.id == TeacherHODSubjectLink.hod_link_id)
        .join(hod_teacher, hod_teacher.id == HODLink.hod_teacher_id)
        .join(hod_user, hod_user.id == hod_teacher.user_id)
        .join(Course, Course.id == HODLink.course_id)
        .join(Branch, Branch.id == HODLink.branch_id)
        .join(Subject, Subject.id == TeacherHODSubjectLink.subject_id)
        .outerjoin(Section, Section.id == TeacherHODSubjectLink.section_id)
        .outerjoin(Class, Class.id == Section.class_id)
    )

    if institution_id:
        q = q.where(HODLink.institution_id == institution_id)
    if course_id:
        q = q.where(HODLink.course_id == course_id)
    if branch_id:
        q = q.where(HODLink.branch_id == branch_id)
    if current_user and not _has_admin_teacher_scope(current_user):
        actor_teacher = await _current_teacher_profile(db, current_user)
        if has_any_role(current_user, {"hod"}):
            q = q.where(HODLink.hod_teacher_id == actor_teacher.id)
        else:
            q = q.where(TeacherHODSubjectLink.teacher_id == actor_teacher.id)

    rows = (await db.execute(q.order_by(teacher_user.full_name.asc(), Subject.name.asc()))).mappings().all()
    return [dict(row) for row in rows]


async def list_institution_timetable(
    db: AsyncSession,
    institution_id: str,
    day_of_week: TimetableDay | None = None,
) -> list[dict]:
    q = (
        select(
            TeacherTimetable.id,
            TeacherTimetable.teacher_id,
            User.full_name.label("teacher_name"),
            Teacher.designation.label("teacher_designation"),
            TeacherTimetable.class_id,
            Class.name.label("class_name"),
            TeacherTimetable.section_id,
            Section.name.label("section_name"),
            TeacherTimetable.subject_id,
            Subject.name.label("subject_name"),
            Branch.id.label("branch_id"),
            Branch.name.label("branch_name"),
            TeacherTimetable.academic_year_id,
            AcademicYear.label.label("academic_year_label"),
            TeacherTimetable.day_of_week,
            TeacherTimetable.start_time,
            TeacherTimetable.end_time,
            TeacherTimetable.room_no,
            TeacherTimetable.version_no,
            TeacherTimetable.is_active,
            literal(None).label("session_id"),
            literal(None).label("session_status"),
            literal(None).label("session_date"),
        )
        .join(Teacher, Teacher.id == TeacherTimetable.teacher_id)
        .join(User, User.id == Teacher.user_id)
        .join(Class, Class.id == TeacherTimetable.class_id)
        .join(Section, Section.id == TeacherTimetable.section_id)
        .join(Subject, Subject.id == TeacherTimetable.subject_id)
        .join(Branch, Branch.id == Class.branch_id)
        .join(AcademicYear, AcademicYear.id == TeacherTimetable.academic_year_id)
        .where(
            User.institution_id == institution_id,
            TeacherTimetable.is_active == True,
        )
    )
    if day_of_week:
        q = q.where(TeacherTimetable.day_of_week == day_of_week)

    rows = (
        await db.execute(
            q.order_by(
                TeacherTimetable.day_of_week.asc(),
                TeacherTimetable.start_time.asc(),
                User.full_name.asc(),
                Class.name.asc(),
                Section.name.asc(),
            )
        )
    ).mappings().all()
    return [dict(row) for row in rows]


async def get_hod_branch_analytics(db: AsyncSession, user_id: str, academic_year_id: str | None = None) -> dict:
    teacher = (
        await db.execute(select(Teacher).where(Teacher.user_id == user_id))
    ).scalar_one_or_none()
    if not teacher:
        raise NotFoundError("Teacher profile not found for current user")

    year = None
    if academic_year_id:
        year = (
            await db.execute(select(AcademicYear).where(AcademicYear.id == academic_year_id))
        ).scalar_one_or_none()
    if not year:
        year = (
            await db.execute(
                select(AcademicYear)
                .join(User, User.institution_id == AcademicYear.institution_id)
                .where(User.id == user_id, AcademicYear.is_current == True)
                .limit(1)
            )
        ).scalar_one_or_none()
    if not year:
        year = (
            await db.execute(select(AcademicYear).order_by(AcademicYear.start_date.desc()).limit(1))
        ).scalar_one_or_none()

    links = (
        await db.execute(
            select(
                HODLink.id.label("hod_link_id"),
                HODLink.institution_id,
                HODLink.course_id,
                Course.name.label("course_name"),
                HODLink.branch_id,
                Branch.name.label("branch_name"),
            )
            .join(Course, Course.id == HODLink.course_id)
            .join(Branch, Branch.id == HODLink.branch_id)
            .where(HODLink.hod_teacher_id == teacher.id)
            .order_by(Course.name.asc(), Branch.name.asc())
        )
    ).mappings().all()

    items = []
    today = date.today()
    for link in links:
        branch_id = link["branch_id"]
        year_id = year.id if year else None
        active_students_q = select(StudentAcademicRecord.student_id).where(
            StudentAcademicRecord.branch_id == branch_id,
            StudentAcademicRecord.exited_at == None,
            StudentAcademicRecord.status == StudentStatus.ACTIVE,
        )
        if year_id:
            active_students_q = active_students_q.where(StudentAcademicRecord.academic_year_id == year_id)

        student_count = (
            await db.execute(select(func.count()).select_from(active_students_q.subquery()))
        ).scalar() or 0

        class_count = (
            await db.execute(select(func.count(func.distinct(Class.id))).where(Class.branch_id == branch_id))
        ).scalar() or 0

        section_count = (
            await db.execute(
                select(func.count(func.distinct(Section.id)))
                .join(Class, Class.id == Section.class_id)
                .where(Class.branch_id == branch_id)
            )
        ).scalar() or 0

        teachers_count = (
            await db.execute(
                select(func.count(func.distinct(TeacherTimetable.teacher_id)))
                .join(Class, Class.id == TeacherTimetable.class_id)
                .where(
                    Class.branch_id == branch_id,
                    TeacherTimetable.is_active == True,
                    *(([TeacherTimetable.academic_year_id == year_id] if year_id else [])),
                )
            )
        ).scalar() or 0

        timetable_slots = (
            await db.execute(
                select(func.count(TeacherTimetable.id))
                .join(Class, Class.id == TeacherTimetable.class_id)
                .where(
                    Class.branch_id == branch_id,
                    TeacherTimetable.is_active == True,
                    *(([TeacherTimetable.academic_year_id == year_id] if year_id else [])),
                )
            )
        ).scalar() or 0

        attendance = (
            await db.execute(
                select(
                    func.count(AttendanceRecord.id).label("total"),
                    func.sum(case((AttendanceRecord.status.in_([AttendanceStatus.PRESENT, AttendanceStatus.LATE]), 1), else_=0)).label("attended"),
                    func.sum(case((AttendanceRecord.status == AttendanceStatus.ABSENT, 1), else_=0)).label("absent"),
                )
                .join(AttendanceSession, AttendanceSession.id == AttendanceRecord.session_id)
                .join(Section, Section.id == AttendanceSession.section_id)
                .join(Class, Class.id == Section.class_id)
                .where(
                    Class.branch_id == branch_id,
                    *(([AttendanceSession.academic_year_id == year_id] if year_id else [])),
                )
            )
        ).first()
        attendance_total = int(attendance.total or 0) if attendance else 0
        attended = int(attendance.attended or 0) if attendance else 0
        absent_total = int(attendance.absent or 0) if attendance else 0

        absent_today = (
            await db.execute(
                select(func.count(AttendanceRecord.id))
                .join(AttendanceSession, AttendanceSession.id == AttendanceRecord.session_id)
                .join(Section, Section.id == AttendanceSession.section_id)
                .join(Class, Class.id == Section.class_id)
                .where(
                    Class.branch_id == branch_id,
                    AttendanceSession.session_date == today,
                    AttendanceRecord.status == AttendanceStatus.ABSENT,
                    *(([AttendanceSession.academic_year_id == year_id] if year_id else [])),
                )
            )
        ).scalar() or 0

        marks_uploaded = (
            await db.execute(
                select(func.count(Mark.id))
                .join(ExamSubject, ExamSubject.id == Mark.exam_subject_id)
                .join(Exam, Exam.id == ExamSubject.exam_id)
                .join(Subject, Subject.id == ExamSubject.subject_id)
                .where(
                    Subject.branch_id == branch_id,
                    *(([Exam.academic_year_id == year_id] if year_id else [])),
                )
            )
        ).scalar() or 0

        items.append(
            {
                "hod_link_id": link["hod_link_id"],
                "institution_id": link["institution_id"],
                "course_id": link["course_id"],
                "course_name": link["course_name"],
                "branch_id": branch_id,
                "branch_name": link["branch_name"],
                "student_count": int(student_count),
                "class_count": int(class_count),
                "section_count": int(section_count),
                "teachers_count": int(teachers_count),
                "timetable_slots": int(timetable_slots),
                "attendance_percentage": round(attended * 100.0 / attendance_total, 2) if attendance_total else 0.0,
                "attendance_records": attendance_total,
                "absent_total": absent_total,
                "absent_today": int(absent_today),
                "marks_uploaded": int(marks_uploaded),
            }
        )

    totals = {
        "branches": len(items),
        "students": sum(row["student_count"] for row in items),
        "classes": sum(row["class_count"] for row in items),
        "sections": sum(row["section_count"] for row in items),
        "teachers": sum(row["teachers_count"] for row in items),
        "timetable_slots": sum(row["timetable_slots"] for row in items),
        "absent_today": sum(row["absent_today"] for row in items),
        "marks_uploaded": sum(row["marks_uploaded"] for row in items),
    }
    totals["attendance_percentage"] = (
        round(sum(row["attendance_percentage"] for row in items) / len(items), 2) if items else 0.0
    )
    return {
        "academic_year_id": str(year.id) if year else None,
        "academic_year_label": year.label if year else None,
        "totals": totals,
        "branches": items,
    }


async def remove_teacher_hod_subject_link(db: AsyncSession, link_id: str) -> None:
    link = (
        await db.execute(select(TeacherHODSubjectLink).where(TeacherHODSubjectLink.id == link_id))
    ).scalar_one_or_none()
    if not link:
        raise NotFoundError("Teacher-HOD subject link not found")
    await db.delete(link)


async def update_teacher_hod_subject_link(
    db: AsyncSession,
    link_id: str,
    teacher_id,
    hod_link_id,
    section_id,
    subject_id,
) -> TeacherHODSubjectLink:
    link = (
        await db.execute(select(TeacherHODSubjectLink).where(TeacherHODSubjectLink.id == link_id))
    ).scalar_one_or_none()
    if not link:
        raise NotFoundError("Teacher-HOD subject link not found")

    teacher = await get_teacher(db, str(teacher_id))
    teacher_user = (await db.execute(select(User).where(User.id == teacher.user_id))).scalar_one()
    hod_link = (await db.execute(select(HODLink).where(HODLink.id == hod_link_id))).scalar_one_or_none()
    if not hod_link:
        raise NotFoundError("HOD link not found")
    if teacher_user.institution_id != hod_link.institution_id:
        raise ValidationError("Teacher does not belong to selected institution")

    subject = (await db.execute(select(Subject).where(Subject.id == subject_id))).scalar_one_or_none()
    if not subject:
        raise NotFoundError("Subject not found")
    if subject.course_id != hod_link.course_id:
        raise ValidationError("Subject does not belong to selected course")
    if subject.branch_id and subject.branch_id != hod_link.branch_id:
        raise ValidationError("Subject does not belong to selected branch")
    section = (
        await db.execute(
            select(Section)
            .join(Class, Class.id == Section.class_id)
            .where(
                Section.id == section_id,
                Class.course_id == hod_link.course_id,
                Class.branch_id == hod_link.branch_id,
            )
        )
    ).scalar_one_or_none()
    if not section:
        raise ValidationError("Section is required and must belong to selected branch/class")

    existing = (
        await db.execute(
            select(TeacherHODSubjectLink).where(
                TeacherHODSubjectLink.teacher_id == teacher.id,
                TeacherHODSubjectLink.hod_link_id == hod_link.id,
                TeacherHODSubjectLink.section_id == section.id,
                TeacherHODSubjectLink.subject_id == subject.id,
                TeacherHODSubjectLink.id != link.id,
            )
        )
    ).scalar_one_or_none()
    if existing:
        raise ConflictError("Teacher is already linked to this class, section, and subject.")

    link.teacher_id = teacher.id
    link.hod_link_id = hod_link.id
    link.section_id = section.id
    link.subject_id = subject.id
    await db.flush()
    await db.refresh(link)
    return link


async def _user_has_teacher_role(db: AsyncSession, user_id: str) -> bool:
    row = (
        await db.execute(
            select(Role.id)
            .join(UserRole, UserRole.role_id == Role.id)
            .where(UserRole.user_id == user_id, _teacher_role_filter())
            .limit(1)
        )
    ).first()
    return row is not None


async def _get_teacher_role(db: AsyncSession, institution_id: str) -> Role | None:
    return (
        await db.execute(
            select(Role).where(
                Role.institution_id == institution_id,
                _teacher_role_filter(),
            )
        )
    ).scalar_one_or_none()


def _teacher_role_filter():
    return or_(
        func.lower(Role.slug) == "teacher",
        func.lower(Role.slug) == "hod",
        func.lower(Role.slug) == "principal",
        func.lower(Role.slug) == "principle",
        func.lower(Role.slug) == "faculty",
        func.lower(Role.slug) == "lecturer",
        func.lower(Role.name) == "teacher",
        func.lower(Role.name) == "hod",
        func.lower(Role.name) == "principal",
        func.lower(Role.name) == "principle",
        func.lower(Role.name) == "faculty",
        func.lower(Role.name) == "lecturer",
        func.lower(Role.name).like("%teacher%"),
        func.lower(Role.name).like("%faculty%"),
        func.lower(Role.name).like("%lecturer%"),
    )


def _read_timetable_rows(filename: str, content: bytes) -> list[dict]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        return [dict(r) for r in reader]
    if lower.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValidationError("XLSX support requires openpyxl. Install it in backend environment.") from exc
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        items: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None and str(v).strip() for v in row):
                continue
            items.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
        return items
    raise ValidationError("Only .csv or .xlsx files are supported")


def _parse_time(value) -> time:
    if isinstance(value, time):
        return value
    value = str(value or "").strip()
    if not value:
        raise ValidationError("Time value is required")
    hh, mm = value.split(":")[:2]
    return time(int(hh), int(mm))


async def _get_timetable_entry(db: AsyncSession, entry_id: str) -> TeacherTimetable:
    entry = (
        await db.execute(select(TeacherTimetable).where(TeacherTimetable.id == entry_id))
    ).scalar_one_or_none()
    if not entry:
        raise NotFoundError("Timetable entry not found")
    return entry


async def _validate_timetable_refs(
    db: AsyncSession,
    teacher: Teacher,
    class_id,
    section_id,
    subject_id,
    academic_year_id,
):
    class_obj = (
        await db.execute(select(Class).where(Class.id == class_id))
    ).scalar_one_or_none()
    if not class_obj:
        raise NotFoundError("Class not found")

    section_obj = (
        await db.execute(select(Section).where(Section.id == section_id))
    ).scalar_one_or_none()
    if not section_obj:
        raise NotFoundError("Section not found")
    if section_obj.class_id != class_obj.id:
        raise ValidationError("Section does not belong to the selected class")

    subject_obj = (
        await db.execute(select(Subject).where(Subject.id == subject_id))
    ).scalar_one_or_none()
    if not subject_obj:
        raise NotFoundError("Subject not found")
    if getattr(subject_obj, "class_id", None):
        if subject_obj.class_id != class_obj.id:
            raise ValidationError("Subject does not belong to the selected class")
    elif subject_obj.branch_id and class_obj.branch_id and subject_obj.branch_id != class_obj.branch_id:
        raise ValidationError("Subject does not belong to the class branch")
    if getattr(subject_obj, "course_id", None) and getattr(class_obj, "course_id", None):
        if subject_obj.course_id != class_obj.course_id:
            raise ValidationError("Subject course does not match selected class course")
    if subject_obj.academic_year_id and subject_obj.academic_year_id != academic_year_id:
        raise ValidationError("Subject does not belong to the selected academic year")
    if class_obj.academic_year_id and class_obj.academic_year_id != academic_year_id:
        raise ValidationError("Class does not belong to the selected academic year")

    class_link = (
        await db.execute(
            select(TeacherClass.id).where(
                TeacherClass.teacher_id == teacher.id,
                TeacherClass.class_id == class_obj.id,
            )
        )
    ).first()
    branch_subject_link = (
        await db.execute(
            select(TeacherHODSubjectLink.id)
            .join(HODLink, HODLink.id == TeacherHODSubjectLink.hod_link_id)
            .where(
                TeacherHODSubjectLink.teacher_id == teacher.id,
                TeacherHODSubjectLink.subject_id == subject_obj.id,
                HODLink.branch_id == class_obj.branch_id,
            )
            .limit(1)
        )
    ).first()
    if not class_link and not branch_subject_link:
        raise ValidationError("Teacher is not linked to the selected class or branch subject")

    return class_obj, section_obj, subject_obj


async def _ensure_timetable_slot_available(
    db: AsyncSession,
    teacher_id,
    section_id,
    day_of_week: TimetableDay,
    start_time,
    end_time=None,
    exclude_id=None,
):
    teacher_filters = [
        TeacherTimetable.teacher_id == teacher_id,
        TeacherTimetable.day_of_week == day_of_week,
    ]
    section_filters = [
        TeacherTimetable.section_id == section_id,
        TeacherTimetable.day_of_week == day_of_week,
    ]
    if exclude_id:
        teacher_filters.append(TeacherTimetable.id != exclude_id)
        section_filters.append(TeacherTimetable.id != exclude_id)
    if end_time:
        teacher_filters.extend(
            [TeacherTimetable.start_time < end_time, TeacherTimetable.end_time > start_time]
        )
        section_filters.extend(
            [TeacherTimetable.start_time < end_time, TeacherTimetable.end_time > start_time]
        )
    else:
        teacher_filters.append(TeacherTimetable.start_time == start_time)
        section_filters.append(TeacherTimetable.start_time == start_time)

    teacher_conflict = (
        await db.execute(select(TeacherTimetable.id).where(*teacher_filters))
    ).first()
    if teacher_conflict:
        raise ConflictError("Teacher already has another timetable entry in this time slot")

    section_conflict = (
        await db.execute(select(TeacherTimetable.id).where(*section_filters))
    ).first()
    if section_conflict:
        raise ConflictError("Section already has another timetable entry in this time slot")


async def _ensure_teacher_subject_link(
    db: AsyncSession,
    teacher_id,
    subject_id,
    section_id,
    academic_year_id,
) -> TeacherSubject:
    link = (
        await db.execute(
            select(TeacherSubject).where(
                TeacherSubject.teacher_id == teacher_id,
                TeacherSubject.subject_id == subject_id,
                TeacherSubject.section_id == section_id,
                TeacherSubject.academic_year_id == academic_year_id,
            )
        )
    ).scalar_one_or_none()
    if link:
        return link

    link = TeacherSubject(
        teacher_id=teacher_id,
        subject_id=subject_id,
        section_id=section_id,
        academic_year_id=academic_year_id,
    )
    db.add(link)
    await db.flush()
    return link


def _validate_time_range(start_time, end_time):
    if end_time <= start_time:
        raise ValidationError("End time must be after start time")


def _date_to_day(value: date | None) -> TimetableDay | None:
    if not value:
        return None
    return TimetableDay(value.strftime("%A").lower())
